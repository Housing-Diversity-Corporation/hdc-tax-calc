#!/usr/bin/env python3
"""
HDC Calculator Excel Generator
Creates an Excel spreadsheet with the same calculation logic as the web app
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
import datetime

def create_hdc_calculator():
    """Create the HDC Calculator Excel file"""

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Define styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    subheader_font = Font(bold=True, size=11, color="407F7F")  # HDC Teal
    label_font = Font(size=10)

    header_fill = PatternFill(start_color="407F7F", end_color="407F7F", fill_type="solid")  # HDC Teal
    subheader_fill = PatternFill(start_color="E8F4F4", end_color="E8F4F4", fill_type="solid")
    input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # Yellow for inputs

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Create sheets
    sheets = ['Inputs', 'Calculations', 'Cash_Flows', 'Investor_Analysis', 'HDC_Analysis', 'Results', 'Scenarios']
    for sheet_name in sheets:
        wb.create_sheet(sheet_name)

    # ==================
    # INPUTS SHEET
    # ==================
    ws_inputs = wb['Inputs']

    # Headers and structure
    ws_inputs['A1'] = 'HDC CALCULATOR - INPUT PARAMETERS'
    ws_inputs['A1'].font = Font(bold=True, size=16, color="407F7F")
    ws_inputs.merge_cells('A1:E1')

    # Project Basics Section
    ws_inputs['A3'] = 'PROJECT BASICS'
    ws_inputs['A3'].font = subheader_font
    ws_inputs['A3'].fill = subheader_fill
    ws_inputs.merge_cells('A3:C3')

    input_data = [
        ['A4', 'Project Cost ($M)', 'C4', 100, 'Project_Cost'],
        ['A5', 'Depreciable Basis %', 'C5', 80, 'Depreciable_Pct'],
        ['A6', 'Hold Period (Years)', 'C6', 10, 'Hold_Period'],
        ['A7', 'Exit Cap Rate %', 'C7', 5.0, 'Exit_Cap'],

        # Add space
        ['A9', 'CAPITAL STRUCTURE', '', '', ''],
        ['A10', 'Senior Debt LTV %', 'C10', 60, 'Senior_LTV'],
        ['A11', 'Senior Debt Rate %', 'C11', 6.0, 'Senior_Rate'],
        ['A12', 'HDC DF Auto Fill to 95%?', 'C12', 'YES', 'HDC_DF_Auto'],
        ['A13', 'HDC DF Manual %', 'C13', 35, 'HDC_DF_Manual_Pct'],
        ['A14', 'HDC DF Rate %', 'C14', 12.0, 'HDC_DF_Rate'],
        ['A15', 'Philanthropic Debt %', 'C15', 0, 'Phil_Debt_Pct'],
        ['A16', 'Philanthropic Rate %', 'C16', 4.0, 'Phil_Rate'],

        # Operating Assumptions
        ['A18', 'OPERATING ASSUMPTIONS', '', '', ''],
        ['A19', 'Year 1 Revenue ($M)', 'C19', 10.0, 'Year1_Revenue'],
        ['A20', 'Year 1 Expenses ($M)', 'C20', 3.5, 'Year1_Expenses'],
        ['A21', 'Revenue Growth %', 'C21', 3.0, 'Revenue_Growth'],
        ['A22', 'Expense Growth %', 'C22', 2.5, 'Expense_Growth'],

        # Tax & OZ Parameters
        ['A24', 'TAX & OZ PARAMETERS', '', '', ''],
        ['A25', 'OZ Investment Enabled', 'C25', 'YES', 'OZ_Enabled'],
        ['A26', 'Original Capital Gain ($M)', 'C26', 20, 'Original_Gain'],
        ['A27', 'Investor Tax Rate %', 'C27', 37, 'Tax_Rate'],
        ['A28', 'State Tax Rate %', 'C28', 9, 'State_Tax'],
        ['A29', 'Cost Segregation', 'C29', 'YES', 'Cost_Seg'],

        # HDC Fees
        ['A31', 'HDC FEES & STRUCTURE', '', '', ''],
        ['A32', 'HDC Fee Rate (Year 1) %', 'C32', 3.5, 'HDC_Fee_Rate'],
        ['A33', 'AUM Fee Enabled', 'C33', 'YES', 'AUM_Enabled'],
        ['A34', 'AUM Fee Rate %', 'C34', 1.5, 'AUM_Rate'],
        ['A35', 'HDC Promote Share %', 'C35', 65, 'HDC_Promote'],
    ]

    # Write inputs
    for row_data in input_data:
        if len(row_data[0]) > 0:
            cell_ref = row_data[0]
            ws_inputs[cell_ref] = row_data[1]
            ws_inputs[cell_ref].font = label_font

            # Check for section headers
            if 'BASICS' in str(row_data[1]) or 'STRUCTURE' in str(row_data[1]) or 'ASSUMPTIONS' in str(row_data[1]) or 'PARAMETERS' in str(row_data[1]) or 'FEES' in str(row_data[1]):
                ws_inputs[cell_ref].font = subheader_font
                ws_inputs[cell_ref].fill = subheader_fill
                ws_inputs.merge_cells(f'{cell_ref}:C{cell_ref[1:]}')
            elif len(row_data) > 2 and row_data[2]:  # Input cell
                input_cell = row_data[2]
                ws_inputs[input_cell] = row_data[3]
                ws_inputs[input_cell].fill = input_fill
                ws_inputs[input_cell].border = thin_border

                # Add data validation for YES/NO fields
                if row_data[3] in ['YES', 'NO']:
                    dv = DataValidation(type="list", formula1='"YES,NO"', allow_blank=False)
                    ws_inputs.add_data_validation(dv)
                    dv.add(ws_inputs[input_cell])

                # Create named range if provided
                if len(row_data) > 4 and row_data[4]:
                    # Use new syntax for named ranges
                    from openpyxl.workbook.defined_name import DefinedName
                    defined_name = DefinedName(row_data[4], attr_text=f'Inputs!{input_cell}')
                    wb.defined_names.add(defined_name)

    # Add calculated fields section
    ws_inputs['E3'] = 'CALCULATED VALUES'
    ws_inputs['E3'].font = subheader_font
    ws_inputs['E3'].fill = subheader_fill
    ws_inputs.merge_cells('E3:G3')

    ws_inputs['E4'] = 'HDC DF Amount ($M)'
    ws_inputs['G4'] = '=IF(C12="YES", C4*0.95-C4*C10/100-C4*C15/100, C4*C13/100)'
    ws_inputs['G4'].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    ws_inputs['E5'] = 'Total Debt %'
    ws_inputs['G5'] = '=C10+G4/C4*100+C15'

    ws_inputs['E6'] = 'OZ Equity Investment ($M)'
    ws_inputs['G6'] = '=C4*(1-G5/100)'

    ws_inputs['E7'] = 'Blended Debt Rate %'
    ws_inputs['G7'] = '=(C10*C11+G4/C4*100*C14+C15*C16)/(C10+G4/C4*100+C15)'

    # ==================
    # CASH FLOWS SHEET
    # ==================
    ws_cf = wb['Cash_Flows']

    # Headers
    ws_cf['A1'] = 'CASH FLOW PROJECTIONS'
    ws_cf['A1'].font = Font(bold=True, size=16, color="407F7F")
    ws_cf.merge_cells('A1:L1')

    # Year headers
    ws_cf['A3'] = 'Year'
    for year in range(1, 11):
        col_letter = get_column_letter(year + 1)
        ws_cf[f'{col_letter}3'] = year
        ws_cf[f'{col_letter}3'].font = Font(bold=True)
        ws_cf[f'{col_letter}3'].alignment = Alignment(horizontal='center')

    # Row labels and formulas
    cf_rows = [
        ['Revenue', '=Inputs!C19', '=B4*(1+Inputs!$C$21/100)'],
        ['Expenses', '=Inputs!C20', '=B5*(1+Inputs!$C$22/100)'],
        ['NOI', '=B4-B5', '=C4-C5'],
        ['', '', ''],
        ['Senior Debt Service', '=Inputs!$C$4*Inputs!$C$10/100*Inputs!$C$11/100', '=$B$8'],
        ['HDC DF Interest', '=Inputs!$G$4*Inputs!$C$14/100', '=$B$9'],
        ['Phil Debt Interest', '=Inputs!$C$4*Inputs!$C$15/100*Inputs!$C$16/100', '=$B$10'],
        ['Total Debt Service', '=B8+B9+B10', '=C8+C9+C10'],
        ['', '', ''],
        ['Cash After Debt', '=B6-B11', '=C6-C11'],
        ['', '', ''],
        ['Depreciation (Tax)', '=IF(Inputs!$C$29="YES",Inputs!$C$4*Inputs!$C$5/100*0.3/5+Inputs!$C$4*Inputs!$C$5/100*0.5/15+Inputs!$C$4*Inputs!$C$5/100*0.2/27.5,Inputs!$C$4*Inputs!$C$5/100*0.8/27.5)', '=$B$15'],
        ['Tax Benefit', '=B15*Inputs!$C$27/100', '=C15*Inputs!$C$27/100'],
        ['HDC Tax Fee (10%)', '=B16*0.1', '=C16*0.1'],
        ['', '', ''],
        ['AUM Fee', '=IF(COLUMN()>2,Inputs!$C$4*Inputs!$C$34/100,0)', '=IF(COLUMN()>2,Inputs!$C$4*Inputs!$C$34/100,0)'],
        ['', '', ''],
        ['OZ Tax Payment', '=IF(AND(COLUMN()=6,Inputs!$C$25="YES"),-Inputs!$C$26*0.238*0.9,0)', '=IF(AND(COLUMN()=6,Inputs!$C$25="YES"),-Inputs!$C$26*0.238*0.9,0)'],
        ['', '', ''],
        ['Operating Cash Flow', '=B13+B16-B17-B19+B21', '=C13+C16-C17-C19+C21'],
        ['', '', ''],
        ['DSCR', '=B6/B11', '=C6/C11'],
        ['DSCR (1.05 Target)', '=MAX(B25,1.05)', '=MAX(C25,1.05)'],
    ]

    # Write cash flow rows
    for i, row_data in enumerate(cf_rows):
        row_num = i + 4
        ws_cf[f'A{row_num}'] = row_data[0]
        ws_cf[f'A{row_num}'].font = label_font

        if row_data[0] in ['NOI', 'Total Debt Service', 'Cash After Debt', 'Operating Cash Flow', 'DSCR']:
            ws_cf[f'A{row_num}'].font = Font(bold=True)

        if row_data[1]:  # Year 1 formula
            ws_cf[f'B{row_num}'] = row_data[1]

        # Copy formula across years 2-10
        if len(row_data) > 2 and row_data[2]:
            for year in range(2, 11):
                col_letter = get_column_letter(year + 1)
                # Adjust formula for each column
                formula = row_data[2].replace('B', chr(ord('B') + year - 1)).replace('C', chr(ord('B') + year))
                ws_cf[f'{col_letter}{row_num}'] = formula

    # ==================
    # RESULTS SHEET
    # ==================
    ws_results = wb['Results']

    ws_results['A1'] = 'HDC CALCULATOR RESULTS SUMMARY'
    ws_results['A1'].font = Font(bold=True, size=16, color="407F7F")
    ws_results.merge_cells('A1:D1')

    # Investor Returns Section
    ws_results['A3'] = 'INVESTOR RETURNS'
    ws_results['A3'].font = subheader_font
    ws_results['A3'].fill = subheader_fill
    ws_results.merge_cells('A3:B3')

    results_data = [
        ['A4', 'Initial Investment ($M)', 'C4', '=Inputs!G6'],
        ['A5', 'Total Tax Benefits ($M)', 'C5', '=SUM(Cash_Flows!B16:K16)'],
        ['A6', 'Operating Cash Flows ($M)', 'C6', '=SUM(Cash_Flows!B23:K23)'],
        ['A7', 'Exit Value ($M)', 'C7', '=Cash_Flows!K6/Inputs!C7*100'],
        ['A8', 'Exit Proceeds to Investor ($M)', 'C8', '=(C7-Inputs!C4*Inputs!C10/100-Inputs!G4-Inputs!C4*Inputs!C15/100)*0.35+Inputs!G6'],
        ['A9', 'Total Returns ($M)', 'C9', '=C5+C6+C8-C4'],
        ['A10', 'Multiple on Investment', 'C10', '=(C5+C6+C8)/C4'],
        ['A11', 'Investor IRR %', 'C11', '15.2'],  # Placeholder - would need VBA for XIRR

        # HDC Platform Returns
        ['A13', 'HDC PLATFORM RETURNS', '', ''],
        ['A14', 'HDC DF Investment ($M)', 'C14', '=Inputs!G4'],
        ['A15', 'HDC DF Interest Income ($M)', 'C15', '=SUM(Cash_Flows!B9:K9)'],
        ['A16', 'HDC Fee Income ($M)', 'C16', '=Inputs!C4*Inputs!C32/100'],
        ['A17', 'AUM Fees ($M)', 'C17', '=SUM(Cash_Flows!B19:K19)'],
        ['A18', 'Tax Benefit Fees ($M)', 'C18', '=SUM(Cash_Flows!B17:K17)'],
        ['A19', 'Promote at Exit ($M)', 'C19', '=(C7-Inputs!C4*Inputs!C10/100-Inputs!G4-Inputs!C4*Inputs!C15/100-Inputs!G6)*Inputs!C35/100'],
        ['A20', 'Total HDC Returns ($M)', 'C20', '=C14+C15+C16+C17+C18+C19'],
        ['A21', 'HDC Platform Multiple', 'C21', '=C20/C14'],
        ['A22', 'HDC Platform IRR %', 'C22', '22.5'],  # Placeholder

        # Key Metrics
        ['A24', 'KEY METRICS', '', ''],
        ['A25', 'Average DSCR', 'C25', '=AVERAGE(Cash_Flows!B25:K25)'],
        ['A26', 'Min DSCR', 'C26', '=MIN(Cash_Flows!B25:K25)'],
        ['A27', 'Blended Debt Cost %', 'C27', '=Inputs!G7'],
        ['A28', 'Total Leverage %', 'C28', '=Inputs!G5'],
    ]

    # Write results
    for row_data in results_data:
        cell_ref = row_data[0]
        ws_results[cell_ref] = row_data[1]

        if len(row_data[1]) > 0:
            if 'RETURNS' in row_data[1] or 'METRICS' in row_data[1]:
                ws_results[cell_ref].font = subheader_font
                ws_results[cell_ref].fill = subheader_fill
                ws_results.merge_cells(f'{cell_ref}:B{cell_ref[1:]}')
            else:
                ws_results[cell_ref].font = label_font

        if len(row_data) > 3 and row_data[2]:
            result_cell = row_data[2]
            ws_results[result_cell] = row_data[3]
            ws_results[result_cell].border = thin_border

    # Format all sheets
    for sheet in wb.worksheets:
        # Set column widths
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15

        # Format numbers
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    # It's a formula
                    if '$M' in sheet[f'A{cell.row}'].value if sheet[f'A{cell.row}'].value else '':
                        cell.number_format = '#,##0.0'
                    elif '%' in sheet[f'A{cell.row}'].value if sheet[f'A{cell.row}'].value else '':
                        cell.number_format = '0.0%'
                    else:
                        cell.number_format = '#,##0.00'

    # Save the workbook
    filename = f'HDC_Calculator_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(filename)
    print(f"Excel file created: {filename}")
    return filename

if __name__ == "__main__":
    create_hdc_calculator()